from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Tuple, List, Optional
from openpyxl.formatting.rule import CellIsRule

import math
import os

import numpy as np
import pandas as pd
import pyomo.environ as pyo

Bucket = Tuple[str, str]               # (zone, family)
Edge = Tuple[str, str, str, str]       # (from_zone, from_family, to_zone, to_family)


# ----------------------------
# Settings / Input structures
# ----------------------------

@dataclass(frozen=True)
class Settings:
    horizon_weeks: int = 8
    lead_time_weeks: int = 0
    integer_hiring: bool = True
    allow_within_zone_transfer: bool = True
    tie_break_transfers: bool = True
    hard_hire_gate: bool = False   # << NEW: ràng buộc cứng cho hiring
    epsilon: float = 1e-6
    big_m: Optional[float] = None



@dataclass
class InputData:
    settings: Settings
    weeks: List[int]
    buckets: List[Bucket]
    allowed_edges: List[Edge]

    A0: Dict[Bucket, float]
    GCOUNT: Dict[Bucket, int]                 # fixed group_count (để report)

    PMC: Dict[Tuple[int, Bucket], float]
    GR: Dict[Tuple[int, Bucket], int]         # group demand (để report)
    UR: Dict[Tuple[int, Bucket], float]       # UR RATE (0..1)

    PD: Dict[Tuple[int, Bucket], float]       # ProductionDemand = PMC*(1-UR)
    PD_per_group: Dict[Tuple[int, Bucket], float]  # PD / GR (report only)

    TO: Dict[Tuple[int, Bucket], float]
    AB: Dict[Tuple[int, Bucket], float]
    ta_cap: Dict[int, float]


# ----------------------------
# Helpers
# ----------------------------

REQUIRED_BUCKET_COLS = ["zone", "family", "hc0"]
REQUIRED_DEMAND_COLS = ["week", "zone", "family", "pmc_demand", "group_demand"]
REQUIRED_TA_COLS = ["week", "hiring_capacity"]


def _finite_max(vals, default: float = 0.0) -> float:
    m = default
    for v in vals:
        try:
            fv = float(v)
        except Exception:
            continue
        if math.isfinite(fv):
            m = max(m, fv)
    return m


def _finite_sum(vals) -> float:
    s = 0.0
    for v in vals:
        try:
            fv = float(v)
        except Exception:
            continue
        if math.isfinite(fv):
            s += fv
    return s


def _to_bool(x) -> bool:
    if isinstance(x, bool):
        return x
    if x is None:
        return False
    if isinstance(x, (int, np.integer)):
        return int(x) != 0
    if isinstance(x, (float, np.floating)):
        if math.isnan(float(x)):
            return False
        return float(x) != 0.0
    s = str(x).strip().lower()
    if s in {"1", "1.0", "true", "yes", "y", "t"}:
        return True
    if s in {"0", "0.0", "false", "no", "n", "f", ""}:
        return False
    return True


def _parse_week(x) -> int:
    if isinstance(x, (int, np.integer)):
        return int(x)
    s = str(x).strip().upper()
    import re
    m = re.match(r"WK\s*(\d+)", s)
    if m:
        return int(m.group(1))
    return int(float(s))


def _safe_int(x, default=0) -> int:
    try:
        if x is None:
            return default
        if isinstance(x, (float, np.floating)) and math.isnan(float(x)):
            return default
        return int(round(float(x)))
    except Exception:
        return default


def _auto_big_m(data: InputData) -> float:
    max_pd = _finite_max(data.PD.values(), default=0.0)
    max_a0 = _finite_max(data.A0.values(), default=0.0)
    sum_cap = _finite_sum(data.ta_cap.get(w, 0.0) for w in data.weeks)
    M = max_pd + max_a0 + sum_cap + 10.0
    if not math.isfinite(M) or M < 0:
        M = 0.0
    return float(M)


# ----------------------------
# Excel IO
# ----------------------------

def read_input_excel(path: str) -> InputData:
    xl = pd.ExcelFile(path)
    sheet_lc = [s.lower() for s in xl.sheet_names]

    # ---- settings
    if "settings" in sheet_lc:
        real = xl.sheet_names[sheet_lc.index("settings")]
        s_df = pd.read_excel(xl, real).dropna(how="all")
        s_df.columns = [c.strip().lower() for c in s_df.columns]

        kv = {}
        if "key" in s_df.columns and "value" in s_df.columns:
            for _, r in s_df.iterrows():
                k = str(r["key"]).strip().lower()
                kv[k] = r["value"]

        eps_raw = kv.get("epsilon", 1e-6)
        try:
            eps_val = float(eps_raw)
            if not math.isfinite(eps_val):
                eps_val = 1e-6
        except Exception:
            eps_val = 1e-6

        big_m_val = None
        if "big_m" in kv:
            try:
                tmp = float(kv["big_m"])
                if math.isfinite(tmp) and tmp >= 0:
                    big_m_val = tmp
            except Exception:
                big_m_val = None

        settings = Settings(
            horizon_weeks=int(kv.get("horizon_weeks", 8)),
            lead_time_weeks=int(kv.get("lead_time_weeks", 0)),
            integer_hiring=_to_bool(kv.get("integer_hiring", True)),
            allow_within_zone_transfer=_to_bool(kv.get("allow_within_zone_transfer", True)),
            tie_break_transfers=_to_bool(kv.get("tie_break_transfers", True)),
            hard_hire_gate=_to_bool(kv.get("hard_hire_gate", False)),  # << NEW
            epsilon=eps_val,
            big_m=big_m_val,
        )

    else:
        settings = Settings()

    weeks = list(range(1, settings.horizon_weeks + 1))

    # ---- buckets
    if "buckets" not in sheet_lc:
        raise ValueError("Missing required sheet: 'buckets'")
    real = xl.sheet_names[sheet_lc.index("buckets")]
    buckets_df = pd.read_excel(xl, real).dropna(how="all")
    buckets_df.columns = [c.strip().lower() for c in buckets_df.columns]
    for c in REQUIRED_BUCKET_COLS:
        if c not in buckets_df.columns:
            raise ValueError(f"Sheet 'buckets' missing required column: {c}")

    # OPTIONAL cols
    if "group_count" not in buckets_df.columns:
        buckets_df["group_count"] = np.nan
    if "turnover_rate" not in buckets_df.columns:
        buckets_df["turnover_rate"] = np.nan
    if "absent_rate" not in buckets_df.columns:
        buckets_df["absent_rate"] = np.nan

    buckets_df["zone"] = buckets_df["zone"].astype(str).str.strip()
    buckets_df["family"] = buckets_df["family"].astype(str).str.strip()
    buckets_df = buckets_df.drop_duplicates(subset=["zone", "family"], keep="first").reset_index(drop=True)

    bucket_list: List[Bucket] = sorted([(r.zone, r.family) for r in buckets_df.itertuples(index=False)])
    if not bucket_list:
        raise ValueError("Sheet 'buckets' is empty after cleaning.")

    A0: Dict[Bucket, float] = {(r.zone, r.family): float(r.hc0) for r in buckets_df.itertuples(index=False)}

    # fixed group_count (>=1)
    GCOUNT: Dict[Bucket, int] = {}
    for r in buckets_df.itertuples(index=False):
        b = (str(r.zone), str(r.family))
        gc = _safe_int(getattr(r, "group_count", 0), default=0)
        GCOUNT[b] = max(1, gc)  # tránh chia 0

    # ---- demand
    if "demand" not in sheet_lc:
        raise ValueError("Missing required sheet: 'demand'")
    real = xl.sheet_names[sheet_lc.index("demand")]
    demand_df = pd.read_excel(xl, real).dropna(how="all")
    demand_df.columns = [c.strip().lower() for c in demand_df.columns]
    for c in REQUIRED_DEMAND_COLS:
        if c not in demand_df.columns:
            raise ValueError(f"Sheet 'demand' missing required column: {c}")

    demand_df["zone"] = demand_df["zone"].astype(str).str.strip()
    demand_df["family"] = demand_df["family"].astype(str).str.strip()
    demand_df["week"] = demand_df["week"].apply(_parse_week)

    if "ur_multiplier" not in demand_df.columns:
        demand_df["ur_multiplier"] = np.nan
    if "ur_rate" not in demand_df.columns:
        demand_df["ur_rate"] = np.nan

    def derive_ur_rate(row) -> float:
        if pd.notna(row.get("ur_rate")):
            ur = float(row["ur_rate"])
        elif pd.notna(row.get("ur_multiplier")):
            ur = 1.0 - float(row["ur_multiplier"])
        else:
            ur = 0.0
        return float(min(1.0, max(0.0, ur)))

    demand_df["ur_rate_final"] = demand_df.apply(derive_ur_rate, axis=1)
    demand_df = demand_df[demand_df["week"].isin(weeks)].copy()

    # fill all (week x bucket) combos
    idx = pd.MultiIndex.from_product([weeks, bucket_list], names=["week", "bucket"])
    tmp = demand_df.copy()
    tmp["bucket"] = list(zip(tmp["zone"], tmp["family"]))
    tmp = tmp.set_index(["week", "bucket"])[["pmc_demand", "group_demand", "ur_rate_final"]].reindex(idx)


    tmp["pmc_demand"] = tmp["pmc_demand"].fillna(0.0)
    tmp["group_demand"] = tmp["group_demand"].fillna(0.0)
    tmp["ur_rate_final"] = tmp["ur_rate_final"].fillna(0.0)

    tmp = tmp.reset_index()
    tmp[["zone", "family"]] = pd.DataFrame(tmp["bucket"].tolist(), index=tmp.index)
    tmp = tmp.drop(columns=["bucket"])

    PMC: Dict[Tuple[int, Bucket], float] = {}
    GR: Dict[Tuple[int, Bucket], int] = {}
    UR: Dict[Tuple[int, Bucket], float] = {}

    PD: Dict[Tuple[int, Bucket], float] = {}
    PDpg: Dict[Tuple[int, Bucket], float] = {}

    for r in tmp.itertuples(index=False):
        b = (str(r.zone), str(r.family))
        t = int(r.week)

        pmc = float(r.pmc_demand)
        gr_int = max(0, _safe_int(r.group_demand, default=0))
        ur_rate = float(r.ur_rate_final)

        PMC[(t, b)] = pmc
        GR[(t, b)] = gr_int
        UR[(t, b)] = ur_rate

        prod_demand = pmc * (1.0 - ur_rate)  # <<< CORE CHANGE
        PD[(t, b)] = prod_demand

        PDpg[(t, b)] = (prod_demand / float(gr_int)) if gr_int > 0 else 0.0

    # ---- TA capacity
    if "ta_capacity" not in sheet_lc:
        raise ValueError("Missing required sheet: 'ta_capacity'")
    real = xl.sheet_names[sheet_lc.index("ta_capacity")]
    ta_df = pd.read_excel(xl, real).dropna(how="all")
    ta_df.columns = [c.strip().lower() for c in ta_df.columns]
    for c in REQUIRED_TA_COLS:
        if c not in ta_df.columns:
            raise ValueError(f"Sheet 'ta_capacity' missing required column: {c}")
    ta_df["week"] = ta_df["week"].apply(_parse_week)
    ta_df = ta_df[ta_df["week"].isin(weeks)].copy()

    ta_cap = {int(r.week): float(r.hiring_capacity) for r in ta_df.itertuples(index=False)}
    for w in weeks:
        ta_cap.setdefault(w, 0.0)

    # ---- rates (TO/Absent)
    if "rates" in sheet_lc:
        real = xl.sheet_names[sheet_lc.index("rates")]
        rates_df = pd.read_excel(xl, real).dropna(how="all")
        rates_df.columns = [c.strip().lower() for c in rates_df.columns]
        req_cols = ["week", "zone", "family", "turnover_rate", "absent_rate"]
        for col in req_cols:
            if col not in rates_df.columns:
                raise ValueError(f"Sheet 'rates' missing required column: {col}")
        rates_df["zone"] = rates_df["zone"].astype(str).str.strip()
        rates_df["family"] = rates_df["family"].astype(str).str.strip()
        rates_df["week"] = rates_df["week"].apply(_parse_week)
        rates_df = rates_df[rates_df["week"].isin(weeks)].copy()
    else:
        rates_df = pd.DataFrame(columns=["week", "zone", "family", "turnover_rate", "absent_rate"])

    # base constants from buckets then override by rates
    base_rows = []
    for w in weeks:
        for (z, f) in bucket_list:
            row = buckets_df[(buckets_df["zone"] == z) & (buckets_df["family"] == f)].iloc[0]
            base_rows.append({
                "week": w,
                "zone": z,
                "family": f,
                "turnover_rate": float(row["turnover_rate"]) if pd.notna(row["turnover_rate"]) else 0.0,
                "absent_rate": float(row["absent_rate"]) if pd.notna(row["absent_rate"]) else 0.0,
            })
    rates_full = pd.DataFrame(base_rows)

    if not rates_df.empty:
        rates_full = rates_full.merge(
            rates_df,
            on=["week", "zone", "family"],
            how="left",
            suffixes=("", "_ovr"),
        )
        for c in ["turnover_rate", "absent_rate"]:
            ovr = c + "_ovr"
            rates_full[c] = np.where(pd.notna(rates_full[ovr]), rates_full[ovr], rates_full[c])
            rates_full.drop(columns=[ovr], inplace=True)

    for c in ["turnover_rate", "absent_rate"]:
        if (rates_full[c] < -1e-9).any() or (rates_full[c] > 1 + 1e-9).any():
            bad = rates_full[(rates_full[c] < 0) | (rates_full[c] > 1)][["week", "zone", "family", c]].head(10)
            raise ValueError(f"Invalid {c} outside [0,1]. Examples:\n{bad}")

    TO = {(int(r.week), (str(r.zone), str(r.family))): float(r.turnover_rate) for r in rates_full.itertuples(index=False)}
    AB = {(int(r.week), (str(r.zone), str(r.family))): float(r.absent_rate) for r in rates_full.itertuples(index=False)}

    # ---- allowed transfer edges
    def allow_all_pairs() -> List[Edge]:
        ed: List[Edge] = []
        for (zi, fi) in bucket_list:
            for (zj, fj) in bucket_list:
                if (zi, fi) == (zj, fj):
                    continue
                if not settings.allow_within_zone_transfer and zi == zj:
                    continue
                ed.append((zi, fi, zj, fj))
        return ed

    allowed_edges: List[Edge] = []
    if "transfer_allowed" in sheet_lc:
        real = xl.sheet_names[sheet_lc.index("transfer_allowed")]
        tr = pd.read_excel(xl, real).dropna(how="all")
        if tr.empty:
            allowed_edges = allow_all_pairs()
        else:
            tr.columns = [c.strip().lower() for c in tr.columns]
            req = ["from_zone", "from_family", "to_zone", "to_family", "allowed"]
            for c in req:
                if c not in tr.columns:
                    raise ValueError(f"Sheet 'transfer_allowed' missing required column: {c}")
            tr["from_zone"] = tr["from_zone"].astype(str).str.strip()
            tr["from_family"] = tr["from_family"].astype(str).str.strip()
            tr["to_zone"] = tr["to_zone"].astype(str).str.strip()
            tr["to_family"] = tr["to_family"].astype(str).str.strip()
            tr["allowed"] = tr["allowed"].apply(_to_bool)

            allowed_set = {(r.from_zone, r.from_family, r.to_zone, r.to_family)
                           for r in tr.itertuples(index=False) if r.allowed}

            for (zi, fi) in bucket_list:
                for (zj, fj) in bucket_list:
                    if (zi, fi) == (zj, fj):
                        continue
                    if not settings.allow_within_zone_transfer and zi == zj:
                        continue
                    if (zi, fi, zj, fj) in allowed_set:
                        allowed_edges.append((zi, fi, zj, fj))
    else:
        allowed_edges = allow_all_pairs()

    return InputData(
        settings=settings,
        weeks=weeks,
        buckets=bucket_list,
        allowed_edges=allowed_edges,
        A0=A0,
        GCOUNT=GCOUNT,
        PMC=PMC,
        GR=GR,
        UR=UR,
        PD=PD,
        PD_per_group=PDpg,
        TO=TO,
        AB=AB,
        ta_cap=ta_cap,
    )


# ----------------------------
# Pyomo Model
# ----------------------------

def build_model(data: InputData) -> pyo.ConcreteModel:
    s = data.settings
    weeks = data.weeks
    buckets = data.buckets
    edges = data.allowed_edges

    M = s.big_m if s.big_m is not None else _auto_big_m(data)
    if not (isinstance(M, (int, float)) and math.isfinite(float(M)) and float(M) >= 0):
        raise ValueError(f"BigM invalid ({M}). Check input or set a finite big_m in 'settings'.")

    # adjacency
    out_edges: Dict[Bucket, List[Edge]] = {b: [] for b in buckets}
    in_edges: Dict[Bucket, List[Edge]] = {b: [] for b in buckets}
    for e in edges:
        out_edges[(e[0], e[1])].append(e)
        in_edges[(e[2], e[3])].append(e)

    m = pyo.ConcreteModel("Phase1_Planning_v4")
    m.T = pyo.Set(initialize=weeks, ordered=True)
    m.B = pyo.Set(initialize=buckets, dimen=2)
    m.E = pyo.Set(initialize=edges, dimen=4)

    m.BigM = pyo.Param(initialize=float(M), within=pyo.NonNegativeReals, mutable=False)

    # Params
    def init_dict(dct, default=0.0):
        def _rule(mm, t, z, f):
            return float(dct.get((int(t), (str(z), str(f))), default))
        return _rule

    def init_dict_int(dct, default=0):
        def _rule(mm, t, z, f):
            return int(dct.get((int(t), (str(z), str(f))), default))
        return _rule

    m.PMC = pyo.Param(m.T, m.B, initialize=init_dict(data.PMC), within=pyo.NonNegativeReals)
    m.GR  = pyo.Param(m.T, m.B, initialize=init_dict_int(data.GR), within=pyo.NonNegativeIntegers)
    m.UR  = pyo.Param(m.T, m.B, initialize=init_dict(data.UR, default=0.0), within=pyo.UnitInterval)

    m.PD  = pyo.Param(m.T, m.B, initialize=init_dict(data.PD), within=pyo.NonNegativeReals)       # required
    m.PDpg = pyo.Param(m.T, m.B, initialize=init_dict(data.PD_per_group), within=pyo.NonNegativeReals)

    m.TO = pyo.Param(m.T, m.B, initialize=init_dict(data.TO), within=pyo.UnitInterval)
    m.AB = pyo.Param(m.T, m.B, initialize=init_dict(data.AB), within=pyo.UnitInterval)

    def A0_init(mm, z, f):
        return float(data.A0[(str(z), str(f))])
    m.A0 = pyo.Param(m.B, initialize=A0_init, within=pyo.NonNegativeReals)

    def GC_init(mm, z, f):
        return int(data.GCOUNT.get((str(z), str(f)), 1))
    m.GC = pyo.Param(m.B, initialize=GC_init, within=pyo.NonNegativeIntegers)

    def Cap_init(mm, t):
        return float(data.ta_cap.get(int(t), 0.0))
    m.Cap = pyo.Param(m.T, initialize=Cap_init, within=pyo.NonNegativeReals)

    # Vars
    hire_domain = pyo.NonNegativeIntegers if s.integer_hiring else pyo.NonNegativeReals
    m.HR = pyo.Var(m.T, m.B, domain=hire_domain)
    # --- Tie-break objective: hire càng trễ càng tốt
    Tmax = max(weeks)

    m.ObjLateHire = pyo.Objective(
        expr=sum((Tmax - int(t)) * sum(m.HR[t, b] for b in m.B) for t in m.T),
        sense=pyo.minimize
    )
    m.ObjLateHire.deactivate()

    m.A  = pyo.Var(m.T, m.B, domain=pyo.NonNegativeReals)
    m.N  = pyo.Var(m.T, m.B, domain=pyo.NonNegativeReals)

    # slack (HC shortage only)
    m.SHC = pyo.Var(m.T, m.B, domain=pyo.NonNegativeReals)

    # transfers
    m.X = pyo.Var(m.T, m.E, domain=pyo.NonNegativeReals)

    m.can_lend = pyo.Var(m.T, m.B, domain=pyo.Binary)
    m.can_borrow = pyo.Var(m.T, m.B, domain=pyo.Binary)
    m.GapPre = pyo.Expression(
        m.T, m.B,
        rule=lambda mm, t, z, f: mm.N[t, (str(z), str(f))] - mm.PD[t, (str(z), str(f))]
    )
    # -----------------
    # Overstaff slack (dư trước transfer) = max(0, GapPre)
    # -----------------
    m.OHC = pyo.Var(m.T, m.B, domain=pyo.NonNegativeReals)

    # OHC >= GapPre  (nếu GapPre âm thì OHC tự về 0)
    m.C_over = pyo.Constraint(
        m.T, m.B,
        rule=lambda mm, t, z, f: mm.OHC[t, (str(z), str(f))] >= mm.GapPre[t, (str(z), str(f))]
    )

    # Objective: minimize total overstaff (để dùng ở stage riêng)
    m.ObjOver = pyo.Objective(
        expr=sum(m.OHC[t, b] for t in m.T for b in m.B),
        sense=pyo.minimize
    )
    m.ObjOver.deactivate()

    # Expressions Out/In and Net after transfer
    def out_expr(mm, t, z, f):
        b = (str(z), str(f))
        return sum(mm.X[t, e] for e in out_edges[b])

    def in_expr(mm, t, z, f):
        b = (str(z), str(f))
        return sum(mm.X[t, e] for e in in_edges[b])

    m.Out = pyo.Expression(m.T, m.B, rule=out_expr)
    m.In  = pyo.Expression(m.T, m.B, rule=in_expr)

    m.NetAfter = pyo.Expression(
        m.T, m.B,
        rule=lambda mm, t, z, f: mm.N[t, (str(z), str(f))] - mm.Out[t, (str(z), str(f))] + mm.In[t, (str(z), str(f))]
    )

    # -----------------
    # Constraints
    # -----------------

    # (C2) Stock-flow with turnover + lead time
    L = int(s.lead_time_weeks)
    min_t = min(weeks)

    def c2_rule(mm, t, z, f):
        t = int(t)
        b = (str(z), str(f))
        prev_A = mm.A0[b] if t == min_t else mm.A[t - 1, b]
        carried = prev_A * (1 - mm.TO[t, b])

        if L <= 0:
            inflow_hire = mm.HR[t, b]
        else:
            inflow_hire = mm.HR[t - L, b] if (t - L) in weeks else 0.0

        return mm.A[t, b] == carried + inflow_hire

    m.C2 = pyo.Constraint(m.T, m.B, rule=c2_rule)

    # (C3) Net after absenteeism
    m.C3 = pyo.Constraint(
        m.T, m.B,
        rule=lambda mm, t, z, f: mm.N[t, (str(z), str(f))] == mm.A[t, (str(z), str(f))] * (1 - mm.AB[t, (str(z), str(f))])
    )
    # ---- Pre-hire (lead_time=0) baseline: carried headcount only (không tính HR[t,b])
    L = int(s.lead_time_weeks)
    min_t = min(weeks)

    if s.hard_hire_gate and L != 0:
        raise ValueError("hard_hire_gate currently supports lead_time_weeks=0 only.")

    def carried_A_rule(mm, t, z, f):
        t = int(t)
        b = (str(z), str(f))
        prev_A = mm.A0[b] if t == min_t else mm.A[t - 1, b]
        return prev_A * (1 - mm.TO[t, b])

    m.A_carried = pyo.Expression(m.T, m.B, rule=carried_A_rule)

    # N_before_hire = carried_A * (1-AB)  (chưa cộng HR[t,b])
    m.N_before_hire = pyo.Expression(
        m.T, m.B,
        rule=lambda mm, t, z, f: mm.A_carried[t, (str(z), str(f))] * (1 - mm.AB[t, (str(z), str(f))])
    )

    # Gap_before_hire = N_before_hire - PD
    m.GapBeforeHire = pyo.Expression(
        m.T, m.B,
        rule=lambda mm, t, z, f: mm.N_before_hire[t, (str(z), str(f))] - mm.PD[t, (str(z), str(f))]
    )
    # Binary: 1 nếu bucket đang thiếu (GapBeforeHire < 0) -> được phép tuyển
    m.need_hire = pyo.Var(m.T, m.B, domain=pyo.Binary)
    # =========================
    # Lookahead pre-hire gate (lead_time=0)
    # =========================
    T_max = max(weeks)

    # set weeks excluding last
    m.T_pre = pyo.Set(initialize=[t for t in weeks if t < T_max], ordered=True)

    # Deficit before hire per bucket: DefBH[t,b] = max(0, -GapBeforeHire[t,b])
    m.DefBH = pyo.Var(m.T, m.B, domain=pyo.NonNegativeReals)

    m.C_defbh_ge = pyo.Constraint(
        m.T, m.B,
        rule=lambda mm, t, z, f: mm.DefBH[t, (str(z), str(f))] >= -mm.GapBeforeHire[t, (str(z), str(f))]
    )
    m.C_defbh_le1 = pyo.Constraint(
        m.T, m.B,
        rule=lambda mm, t, z, f: mm.DefBH[t, (str(z), str(f))] <= mm.BigM * mm.need_hire[t, (str(z), str(f))]
    )
    m.C_defbh_le2 = pyo.Constraint(
        m.T, m.B,
        rule=lambda mm, t, z, f: mm.DefBH[t, (str(z), str(f))] <= -mm.GapBeforeHire[t, (str(z), str(f))] + mm.BigM * (
                    1 - mm.need_hire[t, (str(z), str(f))])
    )

    # Total deficit before hire per week
    m.TotalDefBH = pyo.Expression(m.T, rule=lambda mm, t: sum(mm.DefBH[t, b] for b in mm.B))

    # A bigger M for weekly sums
    m.BigMTotal = pyo.Param(initialize=float(M) * max(1, len(buckets)), within=pyo.NonNegativeReals)

    # OverflowNext[t] = 1 if TotalDefBH[t+1] > Cap[t+1]
    m.OverflowNext = pyo.Var(m.T_pre, domain=pyo.Binary)

    m.C_overflow_lb = pyo.Constraint(
        m.T_pre,
        rule=lambda mm, t: (mm.TotalDefBH[t + 1] - mm.Cap[t + 1]) >= s.epsilon - mm.BigMTotal * (1 - mm.OverflowNext[t])
    )
    m.C_overflow_ub = pyo.Constraint(
        m.T_pre,
        rule=lambda mm, t: (mm.TotalDefBH[t + 1] - mm.Cap[t + 1]) <= 0.0 + mm.BigMTotal * mm.OverflowNext[t]
    )

    # prehire_ok[t,b] = OverflowNext[t] AND need_hire[t+1,b]
    m.prehire_ok = pyo.Var(m.T_pre, m.B, domain=pyo.Binary)

    m.C_prehire1 = pyo.Constraint(m.T_pre, m.B,
                                  rule=lambda mm, t, z, f: mm.prehire_ok[t, (str(z), str(f))] <= mm.OverflowNext[t])
    m.C_prehire2 = pyo.Constraint(m.T_pre, m.B,
                                  rule=lambda mm, t, z, f: mm.prehire_ok[t, (str(z), str(f))] <= mm.need_hire[
                                      t + 1, (str(z), str(f))])
    m.C_prehire3 = pyo.Constraint(
        m.T_pre, m.B,
        rule=lambda mm, t, z, f: mm.prehire_ok[t, (str(z), str(f))] >= mm.OverflowNext[t] + mm.need_hire[
            t + 1, (str(z), str(f))] - 1
    )

    # Hiring allowed:
    # - if thiếu tuần hiện tại: need_hire[t,b] = 1
    # - OR nếu tuần sau overflow cap & bucket tuần sau thiếu: prehire_ok[t,b] = 1
    m.C_hire_gate_now_or_prehire = pyo.Constraint(
        m.T_pre, m.B,
        rule=lambda mm, t, z, f:
        mm.HR[t, (str(z), str(f))] <= mm.BigM * mm.need_hire[t, (str(z), str(f))] + mm.BigM * mm.prehire_ok[
            t, (str(z), str(f))]
    )

    # Last week: chỉ cho hire nếu thiếu tuần đó
    m.C_hire_gate_last = pyo.Constraint(
        pyo.Set(initialize=[T_max]), m.B,
        rule=lambda mm, t, z, f:
        mm.HR[t, (str(z), str(f))] <= mm.BigM * mm.need_hire[t, (str(z), str(f))]
    )

    # Nếu need_hire = 0 => GapBeforeHire >= 0  (không thiếu)
    m.C_hire_gate_lb = pyo.Constraint(
        m.T, m.B,
        rule=lambda mm, t, z, f:
        mm.GapBeforeHire[t, (str(z), str(f))] >= 0.0 - mm.BigM * mm.need_hire[t, (str(z), str(f))]
    )

    # Nếu need_hire = 1 => GapBeforeHire <= -epsilon  (đang thiếu thật)
    m.C_hire_gate_ub = pyo.Constraint(
        m.T, m.B,
        rule=lambda mm, t, z, f:
        mm.GapBeforeHire[t, (str(z), str(f))] <= -s.epsilon + mm.BigM * (1 - mm.need_hire[t, (str(z), str(f))])
    )


    # (C4) TA capacity per week
    m.C4 = pyo.Constraint(m.T, rule=lambda mm, t: sum(mm.HR[t, b] for b in mm.B) <= mm.Cap[t])

    # (C5) Meet production demand after transfers + slack
    m.C5 = pyo.Constraint(
        m.T, m.B,
        rule=lambda mm, t, z, f: mm.NetAfter[t, (str(z), str(f))] + mm.SHC[t, (str(z), str(f))] >= mm.PD[t, (str(z), str(f))]
    )

    # -------- Transfer direction logic (same as v1; based on REQUIRED PD) --------
    # gap_pre = N - PD (before transfer)
    # Không được vừa lend vừa borrow
    m.C60 = pyo.Constraint(
        m.T, m.B,
        rule=lambda mm, t, z, f: mm.can_lend[t, (str(z), str(f))] + mm.can_borrow[t, (str(z), str(f))] <= 1
    )

    # Nếu can_lend = 1 => GAP >= 0.5
    m.C61_lend_th = pyo.Constraint(
        m.T, m.B,
        rule=lambda mm, t, z, f: mm.GapPre[t, (str(z), str(f))] >= 0.5 - mm.BigM * (1 - mm.can_lend[t, (str(z), str(f))])
    )

    # Nếu can_borrow = 1 => GAP <= -0.5
    m.C61_borrow_th = pyo.Constraint(
        m.T, m.B,
        rule=lambda mm, t, z, f: mm.GapPre[t, (str(z), str(f))] <= -0.5 + mm.BigM * (
                    1 - mm.can_borrow[t, (str(z), str(f))])
    )

    # Nếu không bật can_lend/can_borrow thì Out/In phải = 0
    m.C62_out_active = pyo.Constraint(
        m.T, m.B,
        rule=lambda mm, t, z, f: mm.Out[t, (str(z), str(f))] <= mm.BigM * mm.can_lend[t, (str(z), str(f))]
    )
    m.C63_in_active = pyo.Constraint(
        m.T, m.B,
        rule=lambda mm, t, z, f: mm.In[t, (str(z), str(f))] <= mm.BigM * mm.can_borrow[t, (str(z), str(f))]
    )

    # Không cho lend vượt quá surplus, và borrow vượt quá deficit (tightening tốt cho solver)
    m.C64_out_limit = pyo.Constraint(
        m.T, m.B,
        rule=lambda mm, t, z, f: mm.Out[t, (str(z), str(f))] <= mm.GapPre[t, (str(z), str(f))] + mm.BigM * (
                    1 - mm.can_lend[t, (str(z), str(f))])
    )
    m.C64_in_limit = pyo.Constraint(
        m.T, m.B,
        rule=lambda mm, t, z, f: mm.In[t, (str(z), str(f))] <= (-mm.GapPre[t, (str(z), str(f))]) + mm.BigM * (
                    1 - mm.can_borrow[t, (str(z), str(f))])
    )

    # Objective stage 1: minimize HC shortage only
    m.ObjSlack = pyo.Objective(
        expr=sum(m.SHC[t, b] for t in m.T for b in m.B),
        sense=pyo.minimize
    )

    return m


# ----------------------------
# Solve (lexicographic)
# ----------------------------

@dataclass
class SolveResult:
    status: str
    termination_condition: str
    total_shortage: float
    total_hiring: float
    total_transfer: float


def _get_highs_solver():
    try:
        from pyomo.contrib.appsi.solvers import Highs
        return Highs(), "appsi"
    except Exception:
        solver = pyo.SolverFactory("highs")
        return solver, "legacy"


def solve_model(
    m: pyo.ConcreteModel,
    tie_break_transfers: bool = True,   # <-- bật/tắt bằng settings trong Excel
    epsilon: float = 1e-6,
) -> SolveResult:
    solver, mode = _get_highs_solver()

    def _solve():
        if mode == "appsi":
            res = solver.solve(m)
            term = str(getattr(res, "termination_condition", "unknown"))
            status = "ok"
        else:
            res = solver.solve(m, tee=False)
            term = str(res.solver.termination_condition)
            status = str(res.solver.status)
        return status, term

    # ----------------------------
    # Stage 1: min shortage
    # ----------------------------
    m.ObjSlack.activate()
    if hasattr(m, "ObjHire"): m.ObjHire.deactivate()
    if hasattr(m, "ObjOver"): m.ObjOver.deactivate()
    if hasattr(m, "ObjMove"): m.ObjMove.deactivate()

    status, term = _solve()
    total_shortage = float(pyo.value(m.ObjSlack))

    if hasattr(m, "FixShortage"):
        m.del_component(m.FixShortage)
    m.FixShortage = pyo.Constraint(
        expr=sum(m.SHC[t, b] for t in m.T for b in m.B) <= total_shortage + epsilon
    )

    # ----------------------------
    # Stage 2: min total hiring
    # ----------------------------
    if hasattr(m, "ObjHire"):
        m.ObjHire.activate()
    else:
        m.ObjHire = pyo.Objective(
            expr=sum(m.HR[t, b] for t in m.T for b in m.B),
            sense=pyo.minimize
        )

    m.ObjSlack.deactivate()
    if hasattr(m, "ObjOver"): m.ObjOver.deactivate()
    if hasattr(m, "ObjMove"): m.ObjMove.deactivate()

    status, term = _solve()
    total_hiring = float(pyo.value(m.ObjHire))

    if hasattr(m, "FixHire"):
        m.del_component(m.FixHire)
    m.FixHire = pyo.Constraint(
        expr=sum(m.HR[t, b] for t in m.T for b in m.B) <= total_hiring + epsilon
    )

    # ----------------------------
    # ----------------------------
    # Stage 3: hire càng trễ càng tốt (tie-break)
    # ----------------------------
    m.ObjLateHire.activate()
    m.ObjHire.deactivate()
    m.ObjSlack.deactivate()
    if hasattr(m, "ObjOver"): m.ObjOver.deactivate()
    if hasattr(m, "ObjMove"): m.ObjMove.deactivate()

    status, term = _solve()
    late_val = float(pyo.value(m.ObjLateHire))

    if hasattr(m, "FixLate"):
        m.del_component(m.FixLate)
    m.FixLate = pyo.Constraint(expr=m.ObjLateHire.expr <= late_val + epsilon)

    m.ObjLateHire.deactivate()

    # ----------------------------
    # Stage 4: min transfer (sum X) - OPTIONAL (bật/tắt bằng Excel)
    # ----------------------------
    total_transfer = 0.0
    if tie_break_transfers:
        # tạo objective mới cho transfer (tránh dính state khi chạy nhiều lần)
        if hasattr(m, "ObjMove"):
            m.del_component(m.ObjMove)
        m.ObjMove = pyo.Objective(
            expr=sum(m.X[t, e] for t in m.T for e in m.E),
            sense=pyo.minimize
        )

        # deactivate các objective khác
        m.ObjSlack.deactivate()
        if hasattr(m, "ObjHire"): m.ObjHire.deactivate()
        if hasattr(m, "ObjOver"): m.ObjOver.deactivate()

        m.ObjMove.activate()

        status, term = _solve()
        total_transfer = float(pyo.value(m.ObjMove))

    return SolveResult(
        status=status,
        termination_condition=term,
        total_shortage=total_shortage,
        total_hiring=total_hiring,
        total_transfer=total_transfer,
    )


# ----------------------------
# Output
# ----------------------------

def extract_results(data: InputData, m: pyo.ConcreteModel) -> Dict[str, pd.DataFrame]:
    T = list(m.T)
    B = list(m.B)
    min_t = min(T)

    rows = []
    for t in T:
        for (z, f) in B:
            b = (str(z), str(f))
            t_int = int(t)

            # ---- inputs / params
            pmc = float(pyo.value(m.PMC[t, b]))
            ur_rate = float(pyo.value(m.UR[t, b]))          # 0..1
            to_rate = float(pyo.value(m.TO[t, b]))          # 0..1
            ab_rate = float(pyo.value(m.AB[t, b]))          # 0..1

            gr = int(pyo.value(m.GR[t, b]))
            gc = int(pyo.value(m.GC[b]))

            # ---- demand
            pd_prod = float(pyo.value(m.PD[t, b]))          # Production Demand = PMC*(1-UR)
            ppl_per_group = (pd_prod / gr) if gr > 0 else 0.0

            # ---- workforce results
            actual = float(pyo.value(m.A[t, b]))            # after turnover + hires-inflow
            net_pre = float(pyo.value(m.N[t, b]))  # CHƯA tính transfer
            hire = float(pyo.value(m.HR[t, b]))

            # --- raw (để tính toán, KHÔNG làm tròn)
            infl_raw = float(pyo.value(m.In[t, b]))
            outf_raw = float(pyo.value(m.Out[t, b]))
            net_after = net_pre - outf_raw + infl_raw

            # --- display only (để report, có làm tròn)
            infl_disp = int(round(infl_raw))
            outf_disp = int(round(outf_raw))

            # nếu làm tròn ra 0 thì để trống (Excel sẽ blank)
            infl_cell = None if infl_disp == 0 else infl_disp
            outf_cell = None if outf_disp == 0 else outf_disp

            gap_pre = net_pre - pd_prod  # CHƯA transfer
            gap_after = net_after - pd_prod  # ĐÃ transfer

            hc_net_before_hiring_after = net_after - hire  # theo “HC net (đã transfer) - hire”
            hc_net_before_hiring_pre = net_pre - hire  # theo “HC net (chưa transfer) - hire”

            shc = float(pyo.value(m.SHC[t, b]))


            rows.append({
                "Week": t_int,
                "Zone": b[0],
                "Family": b[1],

                "GF": f"{b[0]} | {b[1]}",

                "PMC Demand": pmc,
                "UR": ur_rate,
                "Prod Mult (1-UR)": 1.0 - ur_rate,
                "Production Demand": pd_prod,

                "Group Demand": gr,
                "Group Count": gc,
                "People per Group": ppl_per_group,

                "Actual": actual,
                "Turnover Rate": to_rate,
                "Absent": ab_rate,
                "HC net": net_after,                         # hiển thị HC net = NetAfterTransfer
                "Hiring": hire,

                "Gap": gap_after,
                "HC net before hiring": hc_net_before_hiring_after,

                "In": infl_cell,
                "Out": outf_cell,

                "Slack HC": shc,

                "HC net (no transfer)": net_pre,
                "Gap (no transfer)": gap_pre,
                "HC net before hiring (no transfer)": hc_net_before_hiring_pre,
            })

    df_bucket = pd.DataFrame(rows).sort_values(["Week", "Zone", "Family"]).reset_index(drop=True)

    # Transfers long (giữ như bạn đang làm)
    x_rows = []
    for t in T:
        for e in m.E:
            val_raw = float(pyo.value(m.X[t, e]))
            val_int = int(round(val_raw))  # report only

            if val_int == 0:
                continue  # không ghi lên output nếu làm tròn ra 0

            zi, fi, zj, fj = e
            x_rows.append({
                "Week": int(t),
                "From GF": f"{zi} | {fi}",
                "To GF": f"{zj} | {fj}",
                "Transfer HC": val_int,  # ghi số nguyên
            })
    df_x_long = pd.DataFrame(x_rows, columns=["Week", "From GF", "To GF", "Transfer HC"])

    # KPI đơn giản
    kpi = pd.DataFrame([{
        "total_production_demand": float(df_bucket["Production Demand"].sum()),
        "total_hiring": float(df_bucket["Hiring"].sum()),
        "total_transfer": float(df_x_long["Transfer HC"].sum()) if not df_x_long.empty else 0.0,
        "total_slack_hc": float(df_bucket["Slack HC"].sum()),
    }])

    return {
        "KPI": kpi,
        "Bucket Weekly (Long)": df_bucket,
        "Transfers Long": df_x_long,
    }



def _style_excel(path: str) -> None:
    # Optional: make output prettier (bold header, freeze panes, autofilter, width)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        return

    wb = load_workbook(path)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")  # light blue
    hire_fill = PatternFill("solid", fgColor="92D050")  # xanh
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ws in wb.worksheets:
        if ws.max_row >= 1:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            # style header row
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            from openpyxl.styles import PatternFill
            from openpyxl.utils import get_column_letter
            from openpyxl.formatting.rule import CellIsRule

            hire_fill = PatternFill("solid", fgColor="92D050")

            def _apply_hire_conditional(ws, row_idx: int):
                # giả định cột A là label, từ cột B trở đi là tuần / số liệu
                start_col = 2
                end_col = ws.max_column
                if end_col < start_col:
                    return

                # XÓA fill tĩnh (nếu trước đó bạn đã tô cả hàng)
                for c in range(start_col, end_col + 1):
                    ws.cell(row=row_idx, column=c).fill = PatternFill()

                rng = f"{get_column_letter(start_col)}{row_idx}:{get_column_letter(end_col)}{row_idx}"
                rule = CellIsRule(operator="greaterThan", formula=["0"], fill=hire_fill)
                ws.conditional_formatting.add(rng, rule)

            # ---- áp dụng cho các dòng có label Hiring / Total Hiring
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=1).value
                if v is None:
                    continue
                label = str(v).strip().lower()
                if label in {"hiring", "total hiring"}:
                    _apply_hire_conditional(ws, r)

                    # (tuỳ chọn) giữ ô label (cột A) luôn xanh để nhìn ra hàng Hiring
                    ws.cell(row=r, column=1).fill = hire_fill
                    ws.cell(row=r, column=1).font = Font(bold=True)

            # column widths
            for c in range(1, ws.max_column + 1):
                letter = get_column_letter(c)
                max_len = 0
                for r in range(1, min(ws.max_row, 200) + 1):  # avoid super slow on huge sheets
                    v = ws.cell(row=r, column=c).value
                    if v is None:
                        continue
                    max_len = max(max_len, len(str(v)))
                ws.column_dimensions[letter].width = min(45, max(10, max_len + 2))

    wb.save(path)

def _add_gf_dashboard_sheet(wb, df_bucket: pd.DataFrame) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    name = "GF Dashboard"
    if name in wb.sheetnames:
        wb.remove(wb[name])
    ws = wb.create_sheet(name, 0)

    # Styles
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_header = PatternFill("solid", fgColor="7F7F7F")   # dark grey
    fill_demand = PatternFill("solid", fgColor="F4B183")   # orange/yellow-ish
    fill_actual = PatternFill("solid", fgColor="9DC3E6")   # light blue
    fill_hire   = PatternFill("solid", fgColor="92D050")   # green
    fill_gray   = PatternFill("solid", fgColor="D9D9D9")   # grey
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter

    def apply_green_if_positive(row_idx: int):
        start_col = 2
        end_col = 1 + len(weeks)
        rng = f"{get_column_letter(start_col)}{row_idx}:{get_column_letter(end_col)}{row_idx}"
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="greaterThan", formula=["0"], fill=fill_hire)
        )

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from copy import copy

    FONT_NAME = "Aptos Narrow"

    f_header = Font(name=FONT_NAME, bold=True, color="FFFFFF")
    f_title = Font(name=FONT_NAME, bold=True, size=12)
    f_bold = Font(name=FONT_NAME, bold=True)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left   = Alignment(horizontal="left", vertical="center")

    weeks = sorted(df_bucket["Week"].unique().tolist())
    gfs = df_bucket["GF"].drop_duplicates().tolist()

    # Order GF by Zone then Family (optional)
    gfs = sorted(gfs)
    # Force all cells to Aptos Narrow (keep other font attributes like bold/size/color)
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row_cells:
            nf = copy(cell.font)
            nf.name = FONT_NAME
            cell.font = nf

    # helper to write a row
    def write_row(r, label, values, fill=None, bold=False):
        ws.cell(r, 1, label)
        ws.cell(r, 1).alignment = align_left
        ws.cell(r, 1).border = border
        if fill: ws.cell(r, 1).fill = fill
        if bold: ws.cell(r, 1).font = f_bold
        for j, w in enumerate(weeks, start=2):
            v = values.get(w, 0.0)
            v = 0 if v is None else float(v)
            v = int(round(v))  # <-- làm tròn & convert sang integer

            c = ws.cell(r, j, v)
            c.number_format = "0"  # <-- không hiển thị thập phân

            c.alignment = align_center
            c.border = border
            if fill: c.fill = fill
            if bold: c.font = f_bold

    def write_row_pct(r, label, values, fill=None, bold=False):
        ws.cell(r, 1, label)
        ws.cell(r, 1).alignment = align_left
        ws.cell(r, 1).border = border
        if fill: ws.cell(r, 1).fill = fill
        if bold: ws.cell(r, 1).font = f_bold

        for j, w in enumerate(weeks, start=2):
            v = values.get(w, 0.0)
            v = 0.0 if v is None else float(v)  # giữ dạng 0.15
            c = ws.cell(r, j, v)
            c.number_format = "0%"  # hiển thị 15%
            c.alignment = align_center
            c.border = border
            if fill: c.fill = fill
            if bold: c.font = f_bold

    row = 1
    # Column header widths
    ws.column_dimensions["A"].width = 28
    for j in range(2, 2 + len(weeks)):
        ws.column_dimensions[get_column_letter(j)].width = 12

    for gf in gfs:
        sub = df_bucket[df_bucket["GF"] == gf].set_index("Week")

        # Title row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + len(weeks))
        tcell = ws.cell(row, 1, gf)
        tcell.font = f_title
        tcell.alignment = align_left
        row += 1

        # Week header
        ws.cell(row, 1, "Week").fill = fill_header
        ws.cell(row, 1).font = f_header
        ws.cell(row, 1).alignment = align_center
        ws.cell(row, 1).border = border
        for j, w in enumerate(weeks, start=2):
            c = ws.cell(row, j, w)
            c.fill = fill_header
            c.font = f_header
            c.alignment = align_center
            c.border = border
        row += 1

        # Demand block
        write_row(row, "PMC demand", sub["PMC Demand"].to_dict(), fill=fill_demand, bold=True); row += 1
        write_row_pct(row, "%Under Routing", sub["UR"].to_dict(), fill=fill_demand, bold=True)
        row += 1
        write_row(row, "Production Demand", sub["Production Demand"].to_dict(), fill=fill_demand, bold=True); row += 1

        # Actual block
        write_row(row, "Actual", sub["Actual"].to_dict(), fill=fill_actual, bold=True); row += 1
        write_row_pct(row, "Turn over rate", sub["Turnover Rate"].to_dict(), fill=fill_actual, bold=True)
        row += 1

        write_row_pct(row, "Absent", sub["Absent"].to_dict(), fill=fill_actual, bold=True)
        row += 1

        write_row(row, "HC net", sub["HC net (no transfer)"].to_dict(), fill=fill_actual, bold=True); row += 1

        # Hiring row
        write_row(row, "Hiring", sub["Hiring"].to_dict(), fill=None, bold=True)

        # chỉ tô ô label (cột A)
        ws.cell(row, 1).fill = fill_hire

        # chỉ tô các ô tuần nếu >0
        apply_green_if_positive(row)

        row += 1

        # Gap block
        write_row(row, "Gap", sub["Gap (no transfer)"].to_dict(), fill=fill_gray, bold=True); row += 1
        write_row(row, "HC net before hiring", sub["HC net before hiring (no transfer)"].to_dict(), fill=fill_gray, bold=True); row += 2

    # Summary at bottom
    sum_df = df_bucket.groupby("Week", as_index=True).agg({
        "Production Demand": "sum",
        "Actual": "sum",
        "Hiring": "sum",
        "HC net (no transfer)": "sum",
        "Gap (no transfer)": "sum",
    })

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1 + len(weeks))
    scell = ws.cell(row, 1, "TOTAL (All GFs)")
    scell.font = f_title
    scell.alignment = align_left
    row += 1

    # Week header again
    ws.cell(row, 1, "Week").fill = fill_header
    ws.cell(row, 1).font = f_header
    ws.cell(row, 1).alignment = align_center
    ws.cell(row, 1).border = border
    for j, w in enumerate(weeks, start=2):
        c = ws.cell(row, j, w)
        c.fill = fill_header
        c.font = f_header
        c.alignment = align_center
        c.border = border
    row += 1

    write_row(row, "Total Production Demand", sum_df["Production Demand"].to_dict(), fill=fill_demand, bold=True); row += 1
    write_row(row, "Total Actual", sum_df["Actual"].to_dict(), fill=fill_actual, bold=True); row += 1
    write_row(row, "Total HC net", sum_df["HC net (no transfer)"].to_dict(), fill=fill_actual, bold=True); row += 1
    write_row(row, "Total Hiring", sum_df["Hiring"].to_dict(), fill=None, bold=True)
    ws.cell(row, 1).fill = fill_hire
    apply_green_if_positive(row)
    row += 1

    write_row(row, "Total Gap", sum_df["Gap (no transfer)"].to_dict(), fill=fill_gray, bold=True); row += 1


def _add_transfer_matrix_sheet(wb, df_x_long: pd.DataFrame) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    name = "Transfer Matrix"
    if name in wb.sheetnames:
        wb.remove(wb[name])
    ws = wb.create_sheet(name)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fill_header = PatternFill("solid", fgColor="7F7F7F")
    f_header = Font(bold=True, color="FFFFFF")
    f_title = Font(bold=True, size=12)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    if df_x_long.empty:
        ws["A1"] = "No transfers"
        return

    weeks = sorted(df_x_long["Week"].unique().tolist())
    gfs = sorted(set(df_x_long["From GF"]).union(set(df_x_long["To GF"])))

    row = 1
    for w in weeks:
        block = df_x_long[df_x_long["Week"] == w]

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2 + len(gfs))
        c = ws.cell(row, 1, f"Week {w} - Transfer In/Out Matrix")
        c.font = f_title
        c.alignment = align_left
        row += 1

        # header row
        h = ws.cell(row, 1, "From \\ To")
        h.fill = fill_header
        h.font = f_header
        h.alignment = align_center
        h.border = border

        for j, to_gf in enumerate(gfs, start=2):
            cc = ws.cell(row, j, to_gf)
            cc.fill = fill_header
            cc.font = f_header
            cc.alignment = align_center
            cc.border = border
        row += 1

        # build lookup
        pivot = block.pivot_table(index="From GF", columns="To GF", values="Transfer HC", aggfunc="sum", fill_value=0.0)

        for i, from_gf in enumerate(gfs, start=0):
            rr = row + i
            lc = ws.cell(rr, 1, from_gf)
            lc.alignment = align_left
            lc.border = border

            for j, to_gf in enumerate(gfs, start=2):
                val = float(pivot.loc[from_gf, to_gf]) if (from_gf in pivot.index and to_gf in pivot.columns) else 0.0
                cc = ws.cell(rr, j, val if val != 0 else "")
                cc.number_format = "0.00"
                cc.alignment = align_center
                cc.border = border

        row += len(gfs) + 2  # space between weeks

    ws.column_dimensions["A"].width = 28
    # các cột còn lại auto-ish
    from openpyxl.utils import get_column_letter
    for col in range(2, 2 + len(gfs)):
        ws.column_dimensions[get_column_letter(col)].width = 18

def write_output_excel(frames: Dict[str, pd.DataFrame], output_path: str) -> None:
    # 1) write raw sheets by pandas
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for name, df in frames.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)

    # 2) post-process workbook: add formatted sheets
    from openpyxl import load_workbook
    wb = load_workbook(output_path)

    df_bucket = frames.get("Bucket Weekly (Long)")
    df_transfers = frames.get("Transfers Long")

    if df_bucket is not None and not df_bucket.empty:
        _add_gf_dashboard_sheet(wb, df_bucket)

    if df_transfers is not None:
        _add_transfer_matrix_sheet(wb, df_transfers)

    wb.save(output_path)

    # (optional) bạn có thể giữ _style_excel(output_path) nếu muốn style header cho raw sheets
    # nhưng tránh apply lên dashboard vì nó đã format riêng



def run_from_excel(input_path: str, output_path: str) -> Dict[str, pd.DataFrame]:
    data = read_input_excel(input_path)
    model = build_model(data)
    res = solve_model(model, tie_break_transfers=data.settings.tie_break_transfers, epsilon=data.settings.epsilon)

    frames = extract_results(data, model)
    frames["Solve Info"] = pd.DataFrame([{
        "status": res.status,
        "termination_condition": res.termination_condition,
        "total_slack_hc": res.total_shortage,
        "total_hiring": res.total_hiring,
        "total_transfer": res.total_transfer,
        "big_m": float(pyo.value(model.BigM)),
        "num_edges": len(list(model.E)),
    }])

    write_output_excel(frames, output_path)
    return frames


if __name__ == "__main__":
    import argparse
    import time

    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=False, help="Path to input Excel workbook")
    parser.add_argument("--output", required=False, help="Path to output Excel workbook")
    parser.add_argument("--timestamp_output", action="store_true", help="Add timestamp to output filename (avoid Excel lock)")
    args = parser.parse_args()

    default_input = os.path.join(os.getcwd(), "input_template_v2.xlsx")
    default_output = os.path.join(os.getcwd(), "output2.xlsx")

    input_path = args.input or default_input

    if args.output:
        output_path = args.output
    else:
        if args.timestamp_output:
            base, ext = os.path.splitext(default_output)
            output_path = f"{base}_{int(time.time())}{ext}"
        else:
            output_path = default_output

    print("Running with:")
    print("  input :", input_path)
    print("  output:", output_path)

    run_from_excel(input_path, output_path)
    print(f"Done. Wrote: {output_path}")
